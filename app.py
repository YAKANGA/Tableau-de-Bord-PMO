from __future__ import annotations

import json
import os
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DEFAULT_EXCEL_PATH = BASE_DIR / "modele_pmo.xlsx"

PROJECT_ALIASES = {
    "code": ["n", "code_projet", "id_projet"],
    "name": ["nom_du_projet", "projet"],
    "client": ["client_maitre_d_ouvrage", "client"],
    "country": ["pays"],
    "location": ["localisation"],
    "sector": ["secteur"],
    "pole": ["poles", "pole"],
    "entity": ["filiale_entite", "filiales", "filiale"],
    "pole_manager": ["responsable_pole"],
    "entity_manager": ["dirigeant", "responsable_filiale"],
    "sector_manager": ["responsable_secteur"],
    "director": ["nom_du_directeur", "directeur_de_projet"],
    "manager": ["responsable_projets", "responsable_projet"],
    "amount_ht": ["montant_du_marche_fcfa_ht", "cout_marche_ht", "montant_ht"],
    "amount_ttc": ["montant_du_marche_fcfa_ttc", "cout_marche_ttc", "montant_ttc"],
    "budget": ["budget_previsionnel_fcfa_ttc", "budget"],
    "earned_value": ["valeur_acquise"],
    "expenses": ["depenses", "depenses_reelles", "budget_consomme"],
    "invoiced": ["montant_facture"],
    "collected": ["montant_encaisse"],
    "unpaid": ["impayes"],
    "billing_rate": ["taux_de_facturation"],
    "collection_rate": ["taux_d_encaissement"],
    "length_km": ["longueur_total_itineraires_km", "longueur_km"],
    "duration_months": ["duree_des_travaux_mois", "duree_mois", "duree"],
    "start_date": ["date_de_debut", "debut"],
    "contract_end_date": ["date_de_fin_contractuelle", "fin_contrat"],
    "extended_end_date": ["date_de_fin_prolongee", "fin_prolongee"],
    "planned_progress": ["avancement_prevu", "avancement_prevu_%"],
    "progress": ["avancement_physique", "avancement_physique_%", "avancement_reel"],
    "spi": ["spi"],
    "cpi": ["cpi"],
    "status": ["statut_global_du_projet", "statut"],
    "alert": ["niveau_d_alerte"],
    "risks": ["principaux_risques_blocages", "principaux_risques"],
    "decisions": ["decisions_attendues_de_la_dg", "decisions_attendues"],
    "observations": ["observations"],
    "financing": ["financement"],
    "period": ["periode", "date_periode", "mois"],
}

TEMPLATE_PROJECT_HEADERS = [
    "Code projet",
    "Nom du Projet",
    "Client / Maître d’ouvrage",
    "Pays",
    "Localisation",
    "Secteur",
    "Responsable secteur",
    "Pôle",
    "Responsable pôle",
    "Filiale / Entité",
    "Responsable filiale",
    "Nom du directeur",
    "Responsable projet",
    "Coût marché FCFA HT",
    "Coût marché FCFA TTC",
    "Budget prévisionnel FCFA TTC",
    "Valeur acquise",
    "Dépenses",
    "Montant facturé",
    "Montant encaissé",
    "Impayés",
    "Taux de facturation",
    "Taux d’encaissement",
    "SPI",
    "CPI",
    "% physique",
    "Durée des travaux (mois)",
    "Date de début",
    "Date de fin contractuelle",
    "Date de fin prolongée",
    "Statut du projet",
    "Statut PMO",
    "Principaux risques / blocages",
    "Décisions attendues de la DG",
    "Observations",
]

TEMPLATE_MONTHLY_HEADERS = [
    "Période",
    "Code projet",
    "Avancement prévu %",
    "Avancement réel %",
    "Valeur acquise",
    "Dépenses",
    "Montant facturé",
    "Montant encaissé",
    "Impayés",
    "Taux de facturation",
    "Taux d’encaissement",
    "SPI",
    "CPI",
    "Statut du mois",
    "Commentaires",
]


def normalize_key(value: object) -> str:
    text = str(value or "").strip().lower().replace("’", "'")
    text = "".join(
        char
        for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def display_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def to_number(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\u00a0", "").replace(" ", "").replace(",", ".").replace("%", "")
        try:
            number = float(cleaned)
            return number / 100 if "%" in value else number
        except ValueError:
            return 0.0
    return 0.0


def get_value(row: dict[str, object], field: str, default: object = "") -> object:
    for key in PROJECT_ALIASES[field]:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return default


def get_number(row: dict[str, object], field: str) -> float:
    return to_number(get_value(row, field, 0))


def read_table(ws, header_row: int) -> list[dict[str, object]]:
    headers = [normalize_key(cell.value) for cell in ws[header_row]]
    rows: list[dict[str, object]] = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = [display_value(value) for value in excel_row]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            {
                key: values[index]
                for index, key in enumerate(headers)
                if key and index < len(values)
            }
        )
    return rows


def detect_tables(wb) -> list[tuple[str, int]]:
    tables: list[tuple[str, int]] = []
    for ws in wb.worksheets:
        max_rows = min(ws.max_row or 1, 80)
        for row_index in range(1, max_rows + 1):
            keys = {normalize_key(cell.value) for cell in ws[row_index] if cell.value}
            if {"nom_du_projet", "statut_global_du_projet"}.issubset(keys) or {"projet", "montant_ttc"}.issubset(keys):
                tables.append((ws.title, row_index))
                break
    return tables


def planned_progress(start_value: object, end_value: object, period: date) -> float:
    try:
        start = date.fromisoformat(str(start_value)[:10])
        end = date.fromisoformat(str(end_value)[:10])
    except ValueError:
        return 0.0
    total_days = max((end - start).days, 1)
    elapsed = (period - start).days
    return min(max(elapsed / total_days, 0.0), 1.0)


def has_critical_risk(value: object) -> bool:
    text = normalize_key(value)
    return any(term in text for term in ("critique", "arbitrage", "blocage_majeur", "urgence"))


def classify_pmo_status(spi: float, cpi: float, risk_value: object = "") -> str:
    if spi < 0.8 or cpi < 0.8 or has_critical_risk(risk_value):
        return "Rouge"
    if 0.8 <= spi <= 0.94 or 0.8 <= cpi <= 0.94:
        return "Orange"
    if spi > 0.95 and cpi > 0.95:
        return "Vert"
    return "Orange"


def normalize_project(row: dict[str, object], source_sheet: str, default_period: date) -> dict[str, object] | None:
    name = get_value(row, "name")
    if not name:
        return None

    progress = get_number(row, "progress")
    planned = get_number(row, "planned_progress") or planned_progress(
        get_value(row, "start_date"), get_value(row, "extended_end_date") or get_value(row, "contract_end_date"), default_period
    )
    amount_ht = get_number(row, "amount_ht")
    amount_ttc = get_number(row, "amount_ttc") or amount_ht * 1.18
    budget = get_number(row, "budget")
    earned = get_number(row, "earned_value") or budget * progress
    expenses = get_number(row, "expenses") or budget * progress
    invoiced = get_number(row, "invoiced")
    collected = get_number(row, "collected")
    unpaid = get_number(row, "unpaid") or max(invoiced - collected, 0.0)
    billing_rate = get_number(row, "billing_rate") or (invoiced / amount_ttc if amount_ttc else 0.0)
    collection_rate = get_number(row, "collection_rate") or (collected / invoiced if invoiced else 0.0)
    spi = get_number(row, "spi") or (progress / planned if planned else 0.0)
    cpi = get_number(row, "cpi") or (earned / expenses if expenses else 0.0)
    pmo_status = classify_pmo_status(spi, cpi, get_value(row, "risks"))
    period_value = get_value(row, "period") or default_period.isoformat()[:7]

    return {
        "code": get_value(row, "code") or normalize_key(name).upper()[:16],
        "name": name,
        "client": get_value(row, "client"),
        "country": get_value(row, "country") or "",
        "location": get_value(row, "location") or "Non renseigne",
        "sector": get_value(row, "sector") or "Non renseigne",
        "sectorManager": get_value(row, "sector_manager"),
        "pole": get_value(row, "pole") or infer_pole(source_sheet),
        "poleManager": get_value(row, "pole_manager"),
        "entity": get_value(row, "entity") or source_sheet,
        "entityManager": get_value(row, "entity_manager"),
        "director": get_value(row, "director"),
        "manager": get_value(row, "manager"),
        "amountHt": amount_ht,
        "amountTtc": amount_ttc,
        "budget": budget,
        "earnedValue": earned,
        "expenses": expenses,
        "invoiced": invoiced,
        "collected": collected,
        "unpaid": unpaid,
        "billingRate": billing_rate,
        "collectionRate": collection_rate,
        "spi": spi,
        "cpi": cpi,
        "lengthKm": get_number(row, "length_km"),
        "durationMonths": get_number(row, "duration_months"),
        "startDate": get_value(row, "start_date"),
        "contractEndDate": get_value(row, "contract_end_date"),
        "extendedEndDate": get_value(row, "extended_end_date"),
        "plannedProgress": planned,
        "progress": progress,
        "status": get_value(row, "status") or "Non renseigne",
        "pmoStatus": pmo_status,
        "alert": get_value(row, "alert"),
        "risks": get_value(row, "risks"),
        "decisions": get_value(row, "decisions"),
        "observations": get_value(row, "observations"),
        "financing": get_value(row, "financing"),
        "period": str(period_value)[:7],
        "sourceSheet": source_sheet,
    }


def infer_pole(sheet_name: str) -> str:
    text = sheet_name.upper()
    if "BTP" in text or "ROUTE" in text or "CHANTIER" in text:
        return "BTP"
    if "AXEDECO" in text:
        return "Services"
    return "Non renseigne"


def aggregate(projects: list[dict[str, object]], dimension: str) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for project in projects:
        key = str(project.get(dimension) or "Non renseigne")
        bucket = grouped.setdefault(
            key,
            {
                "name": key,
                "projectCount": 0,
                "amountTtc": 0.0,
                "budget": 0.0,
                "earnedValue": 0.0,
                "expenses": 0.0,
                "progressTotal": 0.0,
                "delayedCount": 0,
                "criticalCount": 0,
            },
        )
        bucket["projectCount"] += 1
        bucket["amountTtc"] += float(project["amountTtc"])
        bucket["budget"] += float(project["budget"])
        bucket["earnedValue"] += float(project["earnedValue"])
        bucket["expenses"] += float(project["expenses"])
        bucket["progressTotal"] += float(project["progress"])
        if "retard" in str(project["status"]).lower():
            bucket["delayedCount"] += 1
        if float(project["spi"]) < 0.8 or float(project["cpi"]) < 0.8:
            bucket["criticalCount"] += 1

    result = []
    for bucket in grouped.values():
        count = int(bucket["projectCount"])
        bucket["averageProgress"] = bucket["progressTotal"] / count if count else 0
        bucket["spi"] = bucket["averageProgress"]
        bucket["cpi"] = bucket["earnedValue"] / bucket["expenses"] if bucket["expenses"] else 0
        del bucket["progressTotal"]
        result.append(bucket)
    return sorted(result, key=lambda item: (-float(item["amountTtc"]), str(item["name"])))


def enrich_project_countries(projects: list[dict[str, object]], references: list[dict[str, object]]) -> None:
    for project in projects:
        if project.get("country"):
            continue
        text = normalize_key(
            " ".join(
                str(project.get(key) or "")
                for key in ("code", "entity", "sourceSheet", "name")
            )
        )
        if "_ci" in f"_{text}_" or text.startswith("ci_") or "cote_d_ivoire" in text:
            project["country"] = "Côte d'Ivoire"
        elif "gabon" in text:
            project["country"] = "Gabon"
        elif "_bj" in f"_{text}_" or "benin" in text:
            project["country"] = "Bénin"
        elif "_sn" in f"_{text}_" or "senegal" in text:
            project["country"] = "Sénégal"
        elif "guinee" in text:
            project["country"] = "Guinée"
        elif "togo" in text:
            project["country"] = "Togo"
        elif "angola" in text:
            project["country"] = "Angola"
        elif "congo" in text:
            project["country"] = "Congo"
        else:
            project["country"] = "Non renseigné"


def load_workbook_data(excel_path: Path) -> dict[str, object]:
    if not excel_path.exists():
        return {
            "ok": False,
            "error": f"Fichier Excel introuvable: {excel_path}",
            "excelPath": str(excel_path),
        }

    default_period = datetime.fromtimestamp(excel_path.stat().st_mtime).date()
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    projects: list[dict[str, object]] = []
    references: list[dict[str, object]] = []

    for sheet_name, header_row in detect_tables(wb):
        ws = wb[sheet_name]
        for row in read_table(ws, header_row):
            project = normalize_project(row, sheet_name, default_period)
            if project:
                projects.append(project)

    if "Données (2)" in wb.sheetnames:
        for row in read_table(wb["Données (2)"], 5):
            if row.get("filiales") or row.get("poles"):
                references.append(
                    {
                        "pole": row.get("poles") or "",
                        "entity": row.get("filiales") or "",
                        "sector": row.get("secteur") or "",
                        "location": row.get("localisation") or "",
                        "status": row.get("statut") or "",
                        "alert": row.get("niveau_d_alerte") or "",
                        "contractType": row.get("type_de_contrat") or "",
                        "manager": row.get("dirigeant") or row.get("directeur_de_projet") or "",
                        "contact": row.get("contact") or "",
                    }
                )

    enrich_project_countries(projects, references)

    periods = sorted({str(project["period"]) for project in projects})
    statuses = Counter(str(project["status"]) for project in projects)
    pmo_statuses = Counter(str(project["pmoStatus"]) for project in projects)
    locations = Counter(str(project["location"]) for project in projects)
    countries = Counter(str(project["country"]) for project in projects)
    sectors = Counter(str(project["sector"]) for project in projects)
    poles = Counter(str(project["pole"]) for project in projects)
    entities = Counter(str(project["entity"]) for project in projects)
    total_amount = sum(float(project["amountTtc"]) for project in projects)
    total_budget = sum(float(project["budget"]) for project in projects)
    total_earned = sum(float(project["earnedValue"]) for project in projects)
    total_expenses = sum(float(project["expenses"]) for project in projects)
    total_invoiced = sum(float(project["invoiced"]) for project in projects)
    total_collected = sum(float(project["collected"]) for project in projects)
    total_unpaid = sum(float(project["unpaid"]) for project in projects)
    average_progress = sum(float(project["progress"]) for project in projects) / len(projects) if projects else 0

    monthly = defaultdict(lambda: {"period": "", "projectCount": 0, "amountTtc": 0.0, "earnedValue": 0.0, "expenses": 0.0, "progressTotal": 0.0})
    for project in projects:
        bucket = monthly[str(project["period"])]
        bucket["period"] = str(project["period"])
        bucket["projectCount"] += 1
        bucket["amountTtc"] += float(project["amountTtc"])
        bucket["earnedValue"] += float(project["earnedValue"])
        bucket["expenses"] += float(project["expenses"])
        bucket["invoiced"] = bucket.get("invoiced", 0.0) + float(project["invoiced"])
        bucket["collected"] = bucket.get("collected", 0.0) + float(project["collected"])
        bucket["unpaid"] = bucket.get("unpaid", 0.0) + float(project["unpaid"])
        bucket["progressTotal"] += float(project["progress"])

    monthly_rows = []
    for bucket in monthly.values():
        count = int(bucket["projectCount"])
        monthly_rows.append(
            {
                "period": bucket["period"],
                "projectCount": count,
                "amountTtc": bucket["amountTtc"],
                "earnedValue": bucket["earnedValue"],
                "expenses": bucket["expenses"],
                "invoiced": bucket.get("invoiced", 0.0),
                "collected": bucket.get("collected", 0.0),
                "unpaid": bucket.get("unpaid", 0.0),
                "averageProgress": bucket["progressTotal"] / count if count else 0,
                "cpi": bucket["earnedValue"] / bucket["expenses"] if bucket["expenses"] else 0,
            }
        )

    return {
        "ok": True,
        "excelPath": str(excel_path),
        "lastModified": datetime.fromtimestamp(excel_path.stat().st_mtime).isoformat(timespec="seconds"),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "summary": {
            "projectCount": len(projects),
            "totalAmountTtc": total_amount,
            "totalBudget": total_budget,
            "totalEarnedValue": total_earned,
            "totalExpenses": total_expenses,
            "totalInvoiced": total_invoiced,
            "totalCollected": total_collected,
            "totalUnpaid": total_unpaid,
            "billingRate": total_invoiced / total_amount if total_amount else 0,
            "collectionRate": total_collected / total_invoiced if total_invoiced else 0,
            "averageProgress": average_progress,
            "portfolioCpi": total_earned / total_expenses if total_expenses else 0,
            "delayedCount": sum(1 for project in projects if "retard" in str(project["status"]).lower()),
            "criticalCount": sum(1 for project in projects if float(project["spi"]) < 0.8 or float(project["cpi"]) < 0.8),
            "statusCounts": dict(statuses),
            "pmoStatusCounts": dict(pmo_statuses),
            "locationCounts": dict(locations),
            "countryCounts": dict(countries),
            "sectorCounts": dict(sectors),
            "poleCounts": dict(poles),
            "entityCounts": dict(entities),
            "periods": periods,
        },
        "projects": projects,
        "references": references,
        "analytics": {
            "byPole": aggregate(projects, "pole"),
            "byEntity": aggregate(projects, "entity"),
            "bySector": aggregate(projects, "sector"),
            "byLocation": aggregate(projects, "location"),
            "monthly": sorted(monthly_rows, key=lambda item: item["period"]),
        },
    }


def style_header(ws, row: int, columns: int) -> None:
    fill = PatternFill("solid", fgColor="17211D")
    for cell in ws[row][:columns]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws, name: str, last_column: str, last_row: int) -> None:
    table = Table(displayName=name, ref=f"A1:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False)
    ws.add_table(table)


def generate_template() -> bytes:
    wb = Workbook()
    ws_ref = wb.active
    ws_ref.title = "Referentiel"
    ws_ref.append(["Pôle", "Responsable pôle", "Filiale / Entité", "Responsable filiale", "Secteur", "Responsable secteur", "Pays", "Localisation", "Type de contrat", "Contact"])
    ws_ref.append(["BTP", "", "PORTEO BTP CI", "", "SECTEUR 1", "", "Côte d'Ivoire", "NORD-EST", "Global et forfait", ""])
    ws_ref.append(["BTP", "", "Porteo BTP Gabon", "", "SECTEUR 1", "", "Gabon", "", "", ""])
    style_header(ws_ref, 1, 10)
    add_table(ws_ref, "ReferentielPMO", "J", 100)

    ws_projects = wb.create_sheet("Projets")
    ws_projects.append(TEMPLATE_PROJECT_HEADERS)
    for index in range(2, 102):
        ws_projects.append([""] * len(TEMPLATE_PROJECT_HEADERS))
        ws_projects[f"O{index}"] = f"=N{index}*1.18"
        ws_projects[f"Q{index}"] = f"=P{index}*Z{index}"
        ws_projects[f"U{index}"] = f"=S{index}-T{index}"
        ws_projects[f"V{index}"] = f'=IFERROR(S{index}/O{index},"")'
        ws_projects[f"W{index}"] = f'=IFERROR(T{index}/S{index},"")'
        ws_projects[f"X{index}"] = f'=IFERROR(Z{index}/MAX(0.0001,((TODAY()-AB{index})/(AC{index}-AB{index}))),"")'
        ws_projects[f"Y{index}"] = f'=IFERROR(Q{index}/R{index},"")'
        ws_projects[f"AF{index}"] = f'=IF(OR(X{index}<0.8,Y{index}<0.8,ISNUMBER(SEARCH("critique",AG{index}))),"Rouge",IF(OR(AND(X{index}>=0.8,X{index}<=0.94),AND(Y{index}>=0.8,Y{index}<=0.94)),"Orange",IF(AND(X{index}>0.95,Y{index}>0.95),"Vert","Orange")))'
    style_header(ws_projects, 1, len(TEMPLATE_PROJECT_HEADERS))
    add_table(ws_projects, "ProjetsPMO", "AI", 101)

    ws_month = wb.create_sheet("Suivi mensuel")
    ws_month.append(TEMPLATE_MONTHLY_HEADERS)
    for index in range(2, 302):
        ws_month.append([""] * len(TEMPLATE_MONTHLY_HEADERS))
        ws_month[f"I{index}"] = f"=G{index}-H{index}"
        ws_month[f"J{index}"] = f'=IFERROR(G{index}/E{index},"")'
        ws_month[f"K{index}"] = f'=IFERROR(H{index}/G{index},"")'
        ws_month[f"L{index}"] = f'=IFERROR(D{index}/C{index},"")'
        ws_month[f"M{index}"] = f'=IFERROR(E{index}/F{index},"")'
    style_header(ws_month, 1, len(TEMPLATE_MONTHLY_HEADERS))
    add_table(ws_month, "SuiviMensuelPMO", "O", 301)

    ws_risks = wb.create_sheet("Risques")
    ws_risks.append(["Date", "Code projet", "Risque / Blocage", "Niveau d’alerte", "Responsable", "Action attendue", "Echéance", "Statut"])
    for _ in range(2, 102):
        ws_risks.append([""] * 8)
    style_header(ws_risks, 1, 8)
    add_table(ws_risks, "RisquesPMO", "H", 101)

    ws_decisions = wb.create_sheet("Decisions")
    ws_decisions.append(["Date", "Code projet", "Décision attendue", "Instance", "Responsable", "Echéance", "Statut", "Commentaire"])
    for _ in range(2, 102):
        ws_decisions.append([""] * 8)
    style_header(ws_decisions, 1, 8)
    add_table(ws_decisions, "DecisionsPMO", "H", 101)

    ws_rules = wb.create_sheet("Regles KPI")
    ws_rules.append(["Indicateur", "Formule"])
    rules = [
        ("Ecart avancement %", "Avancement reel % - Avancement prevu %"),
        ("SPI", "Avancement reel % / Avancement prevu %"),
        ("Valeur acquise", "Budget * Avancement reel %"),
        ("CPI", "Valeur acquise / Depenses"),
        ("Impayes", "Montant facture - Montant encaisse"),
        ("Taux de facturation", "Montant facture / Cout marche TTC"),
        ("Taux d'encaissement", "Montant encaisse / Montant facture"),
    ]
    for rule in rules:
        ws_rules.append(list(rule))
    ws_rules.append([])
    ws_rules.append(["Statut", "Critères d'appréciation"])
    ws_rules.append(["Vert", "Projet maitrise : SPI > 0,95 et CPI > 0,95, sans risque critique identifie"])
    ws_rules.append(["Orange", "Projet sous surveillance : SPI compris entre 0,80 et 0,94 et/ou CPI compris entre 0,80 et 0,94, ou presence de points de vigilance"])
    ws_rules.append(["Rouge", "Projet critique : SPI < 0,80 et/ou CPI < 0,80, ou presence d'un risque critique necessitant un arbitrage"])
    style_header(ws_rules, 1, 2)
    style_header(ws_rules, len(rules) + 3, 2)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = min(max(len(str(cell.value or "")) for cell in column_cells[:20]) + 3, 38)
            ws.column_dimensions[letter].width = max(max_len, 12)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


class PMOHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/dashboard":
            query = parse_qs(parsed.query)
            excel_path = Path(query.get("path", [os.environ.get("PMO_EXCEL_PATH", str(DEFAULT_EXCEL_PATH))])[0])
            self.send_json(load_workbook_data(excel_path))
            return

        if parsed.path == "/api/template":
            self.send_file(generate_template(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "modele_pmo.xlsx")
            return

        requested = "index.html" if parsed.path in ("/", "") else parsed.path.lstrip("/")
        file_path = (STATIC_DIR / requested).resolve()
        if not str(file_path).startswith(str(STATIC_DIR.resolve())) or not file_path.exists():
            self.send_error(404)
            return

        content_type = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "application/javascript; charset=utf-8",
            ".svg": "image/svg+xml",
        }.get(file_path.suffix, "application/octet-stream")
        self.send_bytes(file_path.read_bytes(), content_type, extra_headers={"Cache-Control": "no-store"})

    def send_json(self, payload: dict[str, object]) -> None:
        self.send_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
            200 if payload.get("ok", True) else 404,
            {"Cache-Control": "no-store"},
        )

    def send_file(self, body: bytes, content_type: str, filename: str) -> None:
        self.send_bytes(
            body,
            content_type,
            200,
            {"Content-Disposition": f'attachment; filename="{filename}"', "Cache-Control": "no-store"},
        )

    def send_bytes(self, body: bytes, content_type: str, status: int = 200, extra_headers: dict[str, str] | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        for key, value in (extra_headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args: object) -> None:
        print(f"{self.address_string()} - {format % args}")


def main() -> None:
    host = os.environ.get("PMO_HOST", "127.0.0.1")
    port = int(os.environ.get("PMO_PORT", "8000"))
    server = ThreadingHTTPServer((host, port), PMOHandler)
    print(f"Dashboard PMO disponible sur http://{host}:{port}")
    print(f"Source Excel: {os.environ.get('PMO_EXCEL_PATH', str(DEFAULT_EXCEL_PATH))}")
    server.serve_forever()


if __name__ == "__main__":
    main()
